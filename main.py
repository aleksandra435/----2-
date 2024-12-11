#npm start
import telebot
from telebot import types
from datetime import datetime
from openpyxl import load_workbook

token = "7914110428:AAHawdEUK-84zZWfq0DxBB1SSqeoCOpDvcM"
bot = telebot.TeleBot(token)

# Таблица данных пользователей (id, группа, роль, фио, логин, пароль, предмет)
fn = "bot.xlsx"
wb = load_workbook(fn)
ws = wb['login']
user_data = {}

@bot.message_handler(commands=['start'])
def start(message):
    user_id = str(message.from_user.id)

    # Проверка регистрации пользователя
    def is_registered(user_id):
        for row in ws.iter_rows(values_only=True):
            if row[0] == user_id:
                return row
        return None

    user_row = is_registered(user_id)
    if user_row:
        # Если пользователь уже зарегистрирован, показываем главное меню
        role_info = user_row[2]
        group_info = user_row[3]
        fio_info = user_row[4]
        subject_info = "нет"
        user_data[message.chat.id] = {"id": user_id, "role": role_info, "group": group_info,"fio": fio_info, "fi": subject_info}  # Сохраняем роль в user_data

        show_main_menu(message.chat.id, user_data[message.chat.id])
    else:
        # Начало регистрации нового пользователя
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn1 = types.KeyboardButton("Студент")
        btn2 = types.KeyboardButton("Староста")
        markup.add(btn1, btn2)
        bot.send_message(message.chat.id,f"Здравствуйте, {message.from_user.first_name}! Сначала нужно зарегистрироваться. Выберите роль в меню.",reply_markup=markup)
        # Инициализация временных данных для нового пользователя
        user_data[message.chat.id] = {"id": user_id, "role": None, "group": None, "fio": None, "fi": None}

@bot.message_handler(commands=['help'])
def help_message(message):
    bot.send_message(message.chat.id, ' Здравствуйте! Я бот посещаемости и готов помочь Вам узнать посещаемость группы '
                                      '\nВот что я умею:\n/start - начало работы бота\n/mtuci - официальный '
                                      'сайт\n/raspisanie - расписание занятий \nweek - чёт/нечёт неделя')

@bot.message_handler(commands=['mtuci'])
def mtuci_message(message):
    markup = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton(text="МТУСИ", url="https://mtuci.ru/")
    markup.add(button)
    bot.send_message(message.chat.id, "Более подробно можно изучить тут", reply_markup=markup)

@bot.message_handler(commands=['raspisanie'])
def mtuci_message(message):
    markup = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton(text="Расписание", url="https://mtuci.ru/time-table/")
    markup.add(button)
    bot.send_message(message.chat.id, "Более подробно можно изучить тут", reply_markup=markup)

@bot.message_handler(commands=['week'])
def mtuci_message(message):
    markup = types.InlineKeyboardMarkup()
    button = types.InlineKeyboardButton(text="не/чёт неделя", url="https://whataweek.ru/")
    markup.add(button)
    bot.send_message(message.chat.id, "Узнать какая сейчас неделя можно тут", reply_markup=markup)


@bot.message_handler(content_types=['text'])
def register_user(message):
    # проверка если не то ввели
    user_info = user_data.get(message.chat.id)

    if user_info is None:
        bot.send_message(message.chat.id, "Ошибка! Пожалуйста, начните с команды /start.")
        print(user_info)
        return

    # роль и регистрация(один раз над сделать)
    if message.text in ["Студент", "Староста"]:
        user_info["role"] = message.text
        bot.send_message(message.chat.id, "Теперь введите Ваше ФИО.")
    elif user_info["fio"] is None:
        user_info["fio"] = message.text
        bot.send_message(message.chat.id, "Введите вашу группу (например, БИН2300).")

    elif user_info["role"] == "Староста" and user_info["group"] is None:
        user_info["group"] = message.text
        if user_info["group"] and not is_unique_leader(user_info["group"]):
            bot.send_message(message.chat.id, "В этой группе уже зарегистрирован староста. Пройдите регистрацию заново /start")
        else:
            # Автоматически добавляем старосту в таблицу группы
            group_name = user_info["group"]
            if group_name in wb.sheetnames:
                ws_group = wb[group_name]
            else:
                ws_group = wb.create_sheet(title=group_name)
                ws_group.append(["Предметы:"])
                ws_group.append(["ФИО студентов:"])
            if user_info["fio"] not in [cell.value for cell in ws_group.iter_rows(min_row=3, max_col=1)]:
                ws_group.append([user_info["fio"]])  # Добавляем старосту
                wb.save(fn)
                bot.send_message(message.chat.id, f"Вы добавлены в группу {group_name} как староста.")
            bot.send_message(message.chat.id, "Регистрация старосты завершена. Заполните ФИО  студентов своей группы для дальнейших действий")
            complete_registration(user_info, message.chat.id)
    # группа для студента
    elif user_info["group"] is None:
        user_info["group"] = message.text
        complete_registration(user_info, message.chat.id)
     # добавление студентов старостой
    elif message.text == "Добавить студентов":
        group_name = user_info["group"]
        if group_name in wb.sheetnames:
            ws_group = wb[group_name]
        else:
            ws_group = wb.create_sheet(title=group_name)
            ws_group.append(["Предметы:"])
            ws_group.append(["ФИО студентов:"])
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
        btn_done = types.KeyboardButton("Готово")
        markup.add(btn_done)
        bot.send_message(message.chat.id, "Введите ФИО студентов по одному. Нажмите 'Готово', когда закончите заполнение.", reply_markup=markup)
        user_data[message.chat.id]["adding_students"] = True
        print(f"DEBUG: Бот перешел в режим добавления студентов для группы {group_name}.")
    elif user_data[message.chat.id].get("adding_students"):
        if message.text.lower() == "готово":
            user_data[message.chat.id]["adding_students"] = False
            bot.send_message(message.chat.id, "Добавление студентов завершено.", reply_markup=types.ReplyKeyboardRemove())
            show_main_menu(message.chat.id, user_info)
            print(f"DEBUG: Бот завершил добавление студентов для группы {user_info['group']}.")
        else:
            group_name = user_info["group"]
            ws_group = wb[group_name]
            existing_students = [row[0].value for row in ws_group.iter_rows(min_row=2, max_col=1) if row[0].value]
            if message.text not in existing_students:
                ws_group.append([message.text])
                wb.save(fn)
                bot.send_message(message.chat.id, f"Студент {message.text} добавлен.")
                print(f"DEBUG: Студент {message.text} добавлен в группу {group_name}.")
            else:
                bot.send_message(message.chat.id, f"Студент {message.text} уже существует в списке.")
                print(f"DEBUG: Студент {message.text} уже в списке.")

    # просмотр успеваемости, всем можно
    elif message.text == "Посмотреть успеваемость":
        markup = types.InlineKeyboardMarkup()
        btn_my_site = types.InlineKeyboardButton(text='Наш сайт', url='https://amazing-banoffee-91b1b8.netlify.app/')
        markup.add(btn_my_site)
        bot.send_message(message.chat.id, "Просмотр посещаемости", reply_markup=markup)

    # Заполнение посещаемости по ролям
    elif message.text == "Заполнить успеваемость" and user_info["role"] == "Староста":
        start_attendance_process(message)
    elif message.text == "Заполнить успеваемость" and user_info["role"] != "Староста":
        bot.send_message(message.chat.id, "Если обнаружили ошибку, обратитесь к старосте")


    #заполнение посещаемости (процессы добавления предмета/статус студента/дата в табл)
    elif user_info and "attendance" in user_info:
        attendance_info = user_info["attendance"]

        # Проверка, если пользователь нажал кнопку для добавления нового предмета
        if message.text == "Добавить новый предмет":
            bot.send_message(message.chat.id, "Введите название нового предмета:")
            attendance_info["adding_subject"] = True  # Устанавливаем флаг для добавления предмета
            return  # Возвращаемся, чтобы не обрабатывать другие сообщения

        # Если пользователь в режиме добавления предмета
        elif attendance_info["adding_subject"]:
            # Проверяем, что введено название нового предмета
            new_subject = message.text
            if new_subject.strip():  # Проверяем, не пустой ли введённый предмет
                group_name = user_info["group"]
                ws_group = wb[group_name]

                # Проверяем, существует ли уже такой предмет
                existing_subjects = [cell.value for cell in ws_group[1] if cell.value]
                if new_subject not in existing_subjects:
                    # Добавляем новый предмет в первую строку
                    ws_group.cell(row=1, column=ws_group.max_column + 1, value=new_subject)
                    wb.save(fn)

                    bot.send_message(message.chat.id, f"Предмет '{new_subject}' добавлен.")
                    attendance_info["adding_subject"] = False  # Завершаем процесс добавления

                    # Перезапуск процесса выбора предметов с обновленным списком
                    start_attendance_process(message)
                else:
                    bot.send_message(message.chat.id, "Этот предмет уже существует. Введите другой.")
            else:
                bot.send_message(message.chat.id, "Название предмета не может быть пустым. Попробуйте ещё раз.")

        # Если выбран предмет для посещаемости
        elif attendance_info["subject"] is None:
            # Сохраняем выбранный предмет
            attendance_info["subject"] = message.text
            bot.send_message(message.chat.id, "Введите дату посещаемости (в формате ГГГГ-ММ-ДД):")

        # Если дата посещаемости еще не установлена
        elif attendance_info["date"] is None:
            # Проверяем формат даты
            try:
                attendance_info["date"] = datetime.strptime(message.text, "%Y-%m-%d").date()
                bot.send_message(message.chat.id, "Дата установлена. Теперь отметьте присутствующих студентов.")
                present_students_selection(message.chat.id)
            except ValueError:
                bot.send_message(message.chat.id, "Неверный формат даты. Попробуйте ещё раз (ГГГГ-ММ-ДД).")
        # Завершаем процесс посещаемости
        elif message.text == "Завершить":
            attendance_info["completed"] = True
            save_attendance_data(user_info["group"], attendance_info)
            bot.send_message(message.chat.id, "Посещаемость сохранена.", reply_markup=types.ReplyKeyboardRemove())
            show_main_menu(message.chat.id, user_info)
        # Переключаем статус студента
        elif message.text in attendance_info["students"]:
            student_name = message.text
            # Переключаем статус (Присутствует/Отсутствует)
            attendance_info["students"][student_name] = not attendance_info["students"][student_name]
            status = "Присутствует" if attendance_info["students"][student_name] else "Отсутствует"
            bot.send_message(message.chat.id, f"{student_name}: {status}")
        # Для обработки других вариантов сообщений
        else:
            bot.send_message(message.chat.id, "Неизвестная команда или студент.")

#выбор/сохранение предмета
def start_attendance_process(message):
    user_info = user_data[message.chat.id]
    group_name = user_info["group"]

    if group_name in wb.sheetnames:
        ws_group = wb[group_name]
        all_subjects = [cell.value for cell in ws_group[1] if cell.value]  # Заголовки предметов
        unique_subjects = [subject for subject in set(all_subjects) if subject.lower() != "предметы:"]  # Убираем заголовки

        # Отправляем сообщение с выбором предмета
        markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
        for subject in unique_subjects:
            markup.add(types.KeyboardButton(subject))
        btn_new = types.KeyboardButton("Добавить новый предмет")
        markup.add(btn_new)

        bot.send_message(
            message.chat.id,
            "Выберите предмет и пр/лц для заполнения посещаемости или добавьте новый:",
            reply_markup=markup,
        )

        # Инициализация данных посещаемости
        user_data[message.chat.id]["attendance"] = {
            "subject": None,
            "date": None,
            "students": {},
            "completed": False,
            "adding_subject": False,  # Флаг для добавления предмета
        }
    else:
        bot.send_message(message.chat.id, f"Группа {group_name} не найдена. Заполните свою группу.")

#выбор стутусов студентов по кнопкам
def present_students_selection(chat_id):
    user_info = user_data[chat_id]
    attendance_info = user_info["attendance"]
    group_name = user_info["group"]
    ws_group = wb[group_name]

    # Исключаем заголовки и выбираем только имена студентов
    students = [row[0].value for row in ws_group.iter_rows(min_row=3, max_col=1) if row[0].value]
    if not students:
        bot.send_message(chat_id, "Список студентов пуст. Добавьте студентов в группу.")
        return
    # Инициализация списка студентов
    attendance_info["students"] = {student: False for student in students}  # Инициализация всех студентов как отсутствующих

    # Отправляем кнопки с именами студентов
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for student in students:
        markup.add(types.KeyboardButton(student))
    btn_done = types.KeyboardButton("Завершить")
    markup.add(btn_done)
    bot.send_message(chat_id, "Выберите студентов, которые присутствовали (Чтобы отменить выбор, нажмите на студента ещё раз) . Выберите в меню 'Завершить', когда закончите.", reply_markup=markup)

#сохранение даты и статуса студента
def save_attendance_data(group_name, attendance_info):
    ws_group = wb[group_name]
    date = str(attendance_info["date"])  # Преобразуем дату в строку
    subject = attendance_info["subject"]

    # Добавляем столбец "Процент посещаемости", если его нет
    if "Процент посещаемости" not in [cell.value for cell in ws_group[2]]:
        ws_group.insert_cols(2)
        ws_group.cell(row=2, column=2, value="Процент посещаемости")
        wb.save(fn)
    print(f"DEBUG: Сохраняем посещаемость. Дата: {date}, Предмет: {subject}, Группа: {group_name}")

    if not subject:
        print("DEBUG: Ошибка — предмет не выбран.")
        return "Ошибка: предмет не выбран."

    # Проверяем, существует ли колонка для комбинации предмета и даты
    header_row = 1  # Строка с названием предметов
    date_row = 2    # Строка с датами
    subject_column = None

    for col_idx in range(3, ws_group.max_column + 1):  # Начинаем с 3-й колонки (после ФИО и процента)
        if ws_group.cell(row=header_row, column=col_idx).value == subject and \
                ws_group.cell(row=date_row, column=col_idx).value == date:
            subject_column = col_idx
            break

    # Если колонки нет, добавляем новую колонку
    if not subject_column:
        print(f"DEBUG: Добавляем новую колонку для предмета '{subject}' и даты '{date}'.")
        subject_column = ws_group.max_column + 1
        ws_group.cell(row=header_row, column=subject_column, value=subject)
        ws_group.cell(row=date_row, column=subject_column, value=date)

    # Заполняем данные посещаемости студентов
    print("DEBUG: Заполняем данные посещаемости студентов...")
    for student_idx, (student, present) in enumerate(attendance_info["students"].items(), start=3):
        # Проверяем, существует ли ФИО студента
        if ws_group.cell(row=student_idx, column=1).value != student:
            ws_group.cell(row=student_idx, column=1, value=student)

        # Записываем статус посещаемости
        status = "+" if present else "-"  # Сохраняем '+' или '-' в Excel
        print(f"DEBUG: Записываем статус для студента '{student}': {status}")
        ws_group.cell(row=student_idx, column=subject_column, value=status)

    # Обновляем общий процент посещаемости для каждого студента
    print("DEBUG: Обновляем общий процент посещаемости студентов...")
    for student_idx in range(3, ws_group.max_row + 1):
        total_classes = 0
        attended_classes = 0
        for cell in ws_group[student_idx][2:]:  # Все колонки с посещаемостью (со 2-й колонки)
            if cell.value == "+":
                attended_classes += 1
            if cell.value in ["+", "-"]:
                total_classes += 1
        percentage = (attended_classes / total_classes * 100) if total_classes > 0 else 0
        ws_group.cell(row=student_idx, column=2, value=f"{percentage:.2f}%")  # Процент записывается во второй колонке

    # Удаляем пустые колонки, если появились (перестраховка)
    print("DEBUG: Проверяем наличие пустых колонок...")
    for col_idx in range(ws_group.max_column, 2, -1):  # Проверяем колонки справа налево, начиная с 3-й
        subject_name = ws_group.cell(row=1, column=col_idx).value  # Название предмета в первой строке
        if subject_name:  # Колонка имеет название предмета
            is_empty = all(ws_group.cell(row=row_idx, column=col_idx).value is None
                        for row_idx in range(2, ws_group.max_row + 1))  # Проверяем остальные ячейки
            if is_empty:
                print(f"DEBUG: Удаляем колонку '{subject_name}' в колонке {col_idx}, так как она пустая.")
                ws_group.delete_cols(col_idx)


    # Сохраняем изменения в Excel
    wb.save(fn)
    print(f"DEBUG: Данные успешно сохранены в файл '{fn}'.")
    return "Посещаемость успешно сохранена."



#главное меню
def show_main_menu(chat_id, user_info):
    if "role" not in user_info:
        bot.send_message(chat_id, "Роль пользователя не определена. Пожалуйста, начните регистрацию заново с команды /start.")
        return

    # Формирование меню на основе роли пользователя
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=False)
    btn1 = types.KeyboardButton("Посмотреть успеваемость")
    btn2 = types.KeyboardButton("Заполнить успеваемость")

    if user_info["role"] == "Староста":
        btn3 = types.KeyboardButton("Добавить студентов")
        btn4 = types.KeyboardButton("Логины/пароли группы")
        markup.row(btn3, btn4)  # Кнопки в одной строке
    markup.row(btn1, btn2)  # Кнопки в другой строке

    bot.send_message(chat_id, "Выберите команду в меню.", reply_markup=markup)

#проверка что в группе один староста
def is_unique_leader(group_name):
    # Проверка, есть ли уже староста в группе
    leaders = [row[2] for row in ws.iter_rows(values_only=True) if row[1] == group_name and row[2] == "Староста"]
    return len(leaders) == 0

#сохранение данных регистрации
def complete_registration(user_info, chat_id):
    # Извлечение фамилии из ФИО
    if user_info["fio"]:
        user_info["fi"] = user_info["fio"].split()[0]  # Берём первое слово из ФИО

    # Сохраняем данные пользователя в Excel
    ws.append([
        user_info["id"],
        user_info["group"],
        user_info["role"],
        user_info["fio"],
        user_info["fi"]
    ])
    wb.save(fn)

    # Переход к главному меню
    show_main_menu(chat_id, user_info)



#Есть регистращия и корявое заполнение посещаемости ;)


# Задачи демо и идеи и баги
# отладка если пользователь вместо роли ввел какую-то фигню
#      для старосты прописать кнопку с логинами/паролями группы
# при написании фамалии убрать кнопки студент/староста
#  добавить кнопки назад
#      при добавлении посещаемости повторного предмета меняет данные в первом
# создает лист с фио старосты если какая-то отмена
# что если два раза подряд добавить предет (просто приписывает рядом с другими в одну строку)
# а что если в один день две одинаковых пары
# надо ли разделять на практики/лекции ( или можно записать в предмете это)
# мождо дабывить чет/нечет неделю
# можно добав ссылку на расписание мтуси официальное
# Выбор даты и запись в таб. можно красиво сделать как тут(ввиде календаря) https://habr.com/ru/articles/335886/
# мучиться с сайтом (интерфейсом и как все будет работать (само окно входа и вид сайта тд))

bot.polling(none_stop=True)
