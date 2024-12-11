from flask import Flask, render_template, request
import openpyxl
import os
from urllib.parse import unquote  # Для декодирования URL с кириллицей

# Проверяем наличие файла
print("Файл bot.xlsx существует:", os.path.exists('bot.xlsx'))  # Должно вывести True

app = Flask(__name__)


@app.route('/<group_name>')
def index(group_name):
    try:
        # Декодируем название группы (на случай, если передается кириллица в URL)
        group_name = unquote(group_name)

        # Загружаем файл Excel
        if not os.path.exists('bot.xlsx'):
            raise FileNotFoundError("Файл bot.xlsx не найден.")

        workbook = openpyxl.load_workbook('bot.xlsx')

        # Список всех листов
        sheet_names = workbook.sheetnames
        print("Доступные листы:", sheet_names)  # Вывод всех доступных листов для отладки

        # Проверяем наличие указанной группы
        if group_name not in sheet_names:
            return f"Группа с названием {group_name} не найдена. Доступные группы: {', '.join(sheet_names)}", 404

        # Получаем лист группы
        group_sheet = workbook[group_name]

        # Считываем данные из листа
        data = [[cell.value if cell.value is not None else '' for cell in row] for row in group_sheet.iter_rows()]
        print(f"Данные из группы {group_name}:", data)  # Вывод данных для отладки

        if not data:
            return f"Группа {group_name} найдена, но данные отсутствуют.", 500

        # Передаем данные в шаблон
        return render_template('index.html', data=data, group_name=group_name)

    except FileNotFoundError:
        return "Файл bot.xlsx не найден", 404
    except Exception as e:
        print(f"Ошибка: {e}")  # Вывод полной ошибки в консоль
        return f"Произошла ошибка: {e}", 500


if __name__ == '__main__':
    app.run(debug=True, port=3000)
